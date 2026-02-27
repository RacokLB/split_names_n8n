"""
Integration tests for the POST /process_file endpoint.
Tests XLSX upload, validation, cleaning, and output XLSX generation.
"""
import pytest
import io
import os
from openpyxl import Workbook as OpenpyxlWorkbook


def _create_test_xlsx(rows, sheet_name="Sheet1", header_row=1):
    """
    Helper: creates an in-memory XLSX with the given rows.
    First row of `rows` is treated as headers.
    """
    wb = OpenpyxlWorkbook()
    ws = wb.active
    ws.title = sheet_name

    # Add empty rows before header_row if header_row > 1
    for _ in range(header_row - 1):
        ws.append([])

    for row in rows:
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class TestProcessFileEndpoint:

    def test_valid_file_returns_xlsx(self, client, auth_headers):
        """Upload a valid XLSX → get back a processed XLSX."""
        xlsx = _create_test_xlsx([
            ["CANT. REG", "NAC.", "N° CÉDULA DE IDENTIDAD", "NOMBRE Y APELLIDOS"],
            ["1", "V", 12345678, "CARLOS ANDRES PEREZ GOMEZ"],
            ["2", "E", 7654321, "MARIA DEL PILAR RUIZ"],
        ])
        response = client.post(
            "/process_file?header_row=1",
            headers=auth_headers,
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]
        assert response.headers["x-total-rows"] == "2"
        assert response.headers["x-processed"] == "2"
        assert response.headers["x-rejected"] == "0"

    def test_rejected_rows_counted(self, client, auth_headers):
        """Rows with invalid data are counted as rejected."""
        xlsx = _create_test_xlsx([
            ["CANT. REG", "NAC.", "N° CÉDULA DE IDENTIDAD", "NOMBRE Y APELLIDOS"],
            ["1", "V", 12345678, "JUAN PEREZ"],         # valid
            ["2", "X", "abc", ""],                        # 3 errors: bad nac, bad cedula, empty name
        ])
        response = client.post(
            "/process_file?header_row=1",
            headers=auth_headers,
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.headers["x-processed"] == "1"
        assert response.headers["x-rejected"] == "1"

    def test_empty_rows_skipped(self, client, auth_headers):
        """Completely empty rows are silently skipped."""
        xlsx = _create_test_xlsx([
            ["CANT. REG", "NAC.", "N° CÉDULA DE IDENTIDAD", "NOMBRE Y APELLIDOS"],
            ["1", "V", 12345678, "JUAN PEREZ"],
            [None, None, None, None],  # empty row
            ["3", "E", 7654321, "MARIA LOPEZ"],
        ])
        response = client.post(
            "/process_file?header_row=1",
            headers=auth_headers,
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.headers["x-total-rows"] == "3"
        assert response.headers["x-skipped-empty"] == "1"
        assert response.headers["x-processed"] == "2"

    def test_custom_header_row(self, client, auth_headers):
        """Support custom header row (e.g., row 4 like the real data)."""
        xlsx = _create_test_xlsx(
            rows=[
                ["CANT. REG", "NAC.", "N° CÉDULA DE IDENTIDAD", "NOMBRE Y APELLIDOS"],
                ["1", "V", 12345678, "JUAN PEREZ"],
            ],
            header_row=4,  # headers are on row 4
        )
        response = client.post(
            "/process_file?header_row=4",
            headers=auth_headers,
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.headers["x-processed"] == "1"

    def test_wrong_file_type_returns_400(self, client, auth_headers):
        """Non-XLSX files should be rejected."""
        response = client.post(
            "/process_file?header_row=1",
            headers=auth_headers,
            files={"file": ("test.csv", io.BytesIO(b"a,b,c"), "text/csv")},
        )
        assert response.status_code == 400
        assert "xlsx" in response.json()["detail"].lower()

    def test_missing_auth_returns_403(self, client):
        """File upload without API key should be rejected."""
        xlsx = _create_test_xlsx([
            ["CANT. REG", "NAC.", "N° CÉDULA DE IDENTIDAD", "NOMBRE Y APELLIDOS"],
            ["1", "V", 12345678, "TEST"],
        ])
        response = client.post(
            "/process_file?header_row=1",
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 403

    def test_output_contains_split_names(self, client, auth_headers):
        """Verify the returned XLSX contains the split name columns."""
        xlsx = _create_test_xlsx([
            ["CANT. REG", "NAC.", "N° CÉDULA DE IDENTIDAD", "NOMBRE Y APELLIDOS"],
            ["1", "V", 12345678, "CARLOS ANDRES PEREZ GOMEZ"],
        ])
        response = client.post(
            "/process_file?header_row=1",
            headers=auth_headers,
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200

        # Parse the returned XLSX to verify contents
        from openpyxl import load_workbook
        result_wb = load_workbook(io.BytesIO(response.content))

        # Verify sheets exist
        assert "Processed" in result_wb.sheetnames
        assert "Summary" in result_wb.sheetnames

        # Verify headers in Processed sheet
        ws = result_wb["Processed"]
        headers = [ws.cell(1, col).value for col in range(1, 10)]
        assert "p_nombre" in headers
        assert "p_apellido" in headers
        assert "s_apellido" in headers

        # Verify data row
        assert ws.cell(2, headers.index("p_nombre") + 1).value == "Carlos"
        assert ws.cell(2, headers.index("p_apellido") + 1).value == "Perez"
        assert ws.cell(2, headers.index("s_apellido") + 1).value == "Gomez"

        result_wb.close()
