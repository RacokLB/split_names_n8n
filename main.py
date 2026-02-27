from fastapi import FastAPI, HTTPException, Security, Depends, Request, UploadFile, File
from fastapi.security.api_key import APIKeyHeader
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field
from typing import List
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import os
import io
import re
import logging
from dotenv import load_dotenv
from openpyxl import load_workbook
from xlsxwriter import Workbook

from validators import validate_row, is_row_empty, clean_name

# ── Load .env file (for local development) ─────────────────────────
load_dotenv()

# ── Logging ────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger("name_splitter")

# ── Startup validation ─────────────────────────────────────────────
API_KEY = os.getenv("API_KEY")
if not API_KEY:
    raise RuntimeError(
        "FATAL: API_KEY environment variable is not set. "
        "The server cannot start without a valid API key."
    )

# ── App & rate limiter ─────────────────────────────────────────────
limiter = Limiter(
    key_func=get_remote_address,
    enabled=not os.getenv("TESTING"),  # disable rate limiting during tests
)
app = FastAPI(title="Spanish Name Splitter API")
app.state.limiter = limiter

@app.exception_handler(RateLimitExceeded)
async def rate_limit_handler(request: Request, exc: RateLimitExceeded):
    return JSONResponse(
        status_code=429,
        content={"detail": "Rate limit exceeded. Try again later."},
    )

# ── Authentication ─────────────────────────────────────────────────
API_KEY_NAME = "X-API-Key"
api_key_header = APIKeyHeader(name=API_KEY_NAME, auto_error=True)

async def get_api_key(header_key: str = Security(api_key_header)):
    if header_key == API_KEY:
        return header_key
    else:
        raise HTTPException(status_code=403, detail="Could not validate credentials")

# ── Input validation (Pydantic) ────────────────────────────────────
class NameItem(BaseModel):
    NOMBRE_COMPLETO: str = Field("", max_length=200)
    model_config = {"extra": "allow"}    # preserve passthrough fields (Cedula, N°, etc.)

# ── Core logic ─────────────────────────────────────────────────────
def split_names(full_name: str):
    words = {'de','la','los','del','da','las','o','y','d'}
    original_words = str(full_name).strip().lower().split()

    # group words
    groups = []
    i = 0
    while i < len(original_words):
        chain = []
        while i < len(original_words) and original_words[i] in words:
            chain.append(original_words[i])
            i += 1
        if i < len(original_words):
            word = " ".join(chain) + " " + original_words[i] if chain else original_words[i]
            groups.append(word.strip())
            i += 1
    # Logic for splitting into columns
    num = len(groups)
    answer = {"p_nombre":"", "s_nombre":"","t_nombre":"", "p_apellido":"", "s_apellido":""}

    if num == 2:
        answer["p_nombre"], answer["p_apellido"] = groups[0], groups[1]
    elif num == 3:
        answer["p_nombre"], answer["p_apellido"], answer["s_apellido"] = groups[0], groups[1], groups[2]
    elif num >= 4:
        answer["p_apellido"], answer["s_apellido"] = groups[-2:]
        answer["p_nombre"] = groups[0]
        answer["s_nombre"] = groups[1] if num > 3 else ""
    
    return {k: v.title() for k, v in answer.items() }

# ── Endpoints ──────────────────────────────────────────────────────
@app.get("/")
async def root():
    return {"message": "Name Splitter API is Online"}

@app.post("/split_names")
@limiter.limit("30/minute")
async def process_data(
    request: Request,
    items: List[NameItem],
    api_key: str = Depends(get_api_key),
):
    if len(items) > 20000:
        raise HTTPException(status_code=400, detail="Maximum 20,000 items per request.")

    processed_list = []
    for item in items:
        full_name = item.NOMBRE_COMPLETO
        split_result = split_names(full_name)

        # Merge split names back into the original dictionary
        # .model_dump() preserves all extra passthrough fields
        processed_list.append({**item.model_dump(), **split_result})
    
    return processed_list

# ── Column mapping (raw XLSX headers → internal names) ─────────────
COLUMN_MAP = {
    "CANT. REG":                "N°",
    "NAC.":                     "Nacionalidad",
    "N° CÉDULA DE IDENTIDAD":   "Cedula",
    "NOMBRE Y APELLIDOS":       "NOMBRE_COMPLETO",
}

# Output column order in the result XLSX
OUTPUT_COLUMNS = [
    "N°", "Nacionalidad", "Cedula", "NOMBRE_COMPLETO",
    "p_nombre", "s_nombre", "t_nombre", "p_apellido", "s_apellido",
]


@app.post("/process_file")
@limiter.limit("5/minute")
async def process_file(
    request: Request,
    file: UploadFile = File(...),
    sheet_name: str | None = None,
    header_row: int = 4,
    api_key: str = Depends(get_api_key),
):
    """
    Upload an XLSX file → validate, clean, split names → download processed XLSX.

    - **file**: The XLSX file to process.
    - **sheet_name**: Sheet name to read (default: first sheet).
    - **header_row**: Row number where the column headers are (1-indexed, default 4).
    """
    # ── Validate file type ─────────────────────────────────────────
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx files are accepted."
        )

    # ── Read the XLSX ──────────────────────────────────────────────
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Failed to read XLSX: {e}")
        raise HTTPException(status_code=400, detail=f"Invalid XLSX file: {e}")

    # Select sheet
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise HTTPException(
                status_code=400,
                detail=f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    logger.info(f"Processing sheet: '{ws.title}', header_row={header_row}")

    # ── Extract headers from the specified header row ──────────────
    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    if not header_cells:
        raise HTTPException(status_code=400, detail=f"Row {header_row} is empty — cannot find headers.")

    raw_headers = [str(h).strip() if h else "" for h in header_cells[0]]
    logger.info(f"Raw headers found: {raw_headers}")

    # Map raw headers → internal names
    mapped_headers = []
    for h in raw_headers:
        mapped_headers.append(COLUMN_MAP.get(h, h))

    # ── Process rows ───────────────────────────────────────────────
    processed_rows = []
    rejected_rows = []
    skipped_empty = 0
    total_rows = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        total_rows += 1
        row_values = list(row)

        # Skip entirely empty rows
        if is_row_empty(row_values):
            skipped_empty += 1
            continue

        # Build dict with mapped column names
        row_dict = {}
        for idx, val in enumerate(row_values):
            if idx < len(mapped_headers):
                row_dict[mapped_headers[idx]] = val

        # Validate and clean
        is_valid, cleaned, errors = validate_row(row_dict)

        if is_valid:
            # Split names
            split_result = split_names(cleaned["NOMBRE_COMPLETO"])
            final_row = {**cleaned, **split_result}
            processed_rows.append(final_row)
        else:
            rejected_rows.append({
                "row_number": header_row + total_rows,
                "errors": errors,
                "original_data": row_dict,
            })

    wb.close()

    logger.info(
        f"Done: {len(processed_rows)} processed, "
        f"{len(rejected_rows)} rejected, "
        f"{skipped_empty} empty rows skipped"
    )

    # ── Build output XLSX ──────────────────────────────────────────
    output_buffer = io.BytesIO()
    out_wb = Workbook(output_buffer, {"in_memory": True})

    # Sheet 1: Processed data
    ws_ok = out_wb.add_worksheet("Processed")
    header_fmt = out_wb.add_format({"bold": True, "bg_color": "#4472C4", "font_color": "white"})

    for col_idx, col_name in enumerate(OUTPUT_COLUMNS):
        ws_ok.write(0, col_idx, col_name, header_fmt)

    for row_idx, row_data in enumerate(processed_rows, start=1):
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS):
            ws_ok.write(row_idx, col_idx, row_data.get(col_name, ""))

    # Auto-fit column widths (approximate)
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS):
        ws_ok.set_column(col_idx, col_idx, max(len(col_name) + 2, 15))

    # Sheet 2: Rejected data (for review)
    if rejected_rows:
        ws_rej = out_wb.add_worksheet("Rejected")
        rej_headers = ["Row #", "Errors", "N°", "Nacionalidad", "Cedula", "NOMBRE_COMPLETO"]
        err_fmt = out_wb.add_format({"bold": True, "bg_color": "#FF6B6B", "font_color": "white"})

        for col_idx, h in enumerate(rej_headers):
            ws_rej.write(0, col_idx, h, err_fmt)

        for row_idx, rej in enumerate(rejected_rows, start=1):
            ws_rej.write(row_idx, 0, rej["row_number"])
            ws_rej.write(row_idx, 1, " | ".join(rej["errors"]))
            orig = rej["original_data"]
            ws_rej.write(row_idx, 2, str(orig.get("N°", "")))
            ws_rej.write(row_idx, 3, str(orig.get("Nacionalidad", "")))
            ws_rej.write(row_idx, 4, str(orig.get("Cedula", "")))
            ws_rej.write(row_idx, 5, str(orig.get("NOMBRE_COMPLETO", "")))

        for col_idx, h in enumerate(rej_headers):
            ws_rej.set_column(col_idx, col_idx, max(len(h) + 2, 20))

    # Sheet 3: Summary
    ws_sum = out_wb.add_worksheet("Summary")
    summary_data = [
        ("Total rows read", total_rows),
        ("Empty rows skipped", skipped_empty),
        ("Rows processed successfully", len(processed_rows)),
        ("Rows rejected (see Rejected sheet)", len(rejected_rows)),
    ]
    title_fmt = out_wb.add_format({"bold": True})
    for row_idx, (label, value) in enumerate(summary_data):
        ws_sum.write(row_idx, 0, label, title_fmt)
        ws_sum.write(row_idx, 1, value)
    ws_sum.set_column(0, 0, 40)

    out_wb.close()

    # ── Return the XLSX as a download ──────────────────────────────
    output_buffer.seek(0)
    output_filename = f"processed_{file.filename}"

    return StreamingResponse(
        output_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{output_filename}"',
            "X-Total-Rows": str(total_rows),
            "X-Processed": str(len(processed_rows)),
            "X-Rejected": str(len(rejected_rows)),
            "X-Skipped-Empty": str(skipped_empty),
        },
    )